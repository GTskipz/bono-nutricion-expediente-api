from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from sqlalchemy.orm import Session
from sqlalchemy import text
from uuid import UUID
from datetime import datetime, date
import hashlib
import json
from io import BytesIO

import openpyxl
import re
import unicodedata

from app.core.db import get_db

# ✅ Reusar creación oficial de expediente
from app.routers.expedientes import crear_expediente_core
from app.schemas.expediente import ExpedienteCreate, InfoGeneralIn

router = APIRouter(prefix="/sesan", tags=["SESAN"])


# =====================================================
# Helpers
# =====================================================

def _norm_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def _to_int(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _sha256_bytes(b: bytes) -> str:
    h = hashlib.sha256()
    h.update(b)
    return h.hexdigest()


def _to_cui(v):
    """
    Excel a veces trae el CUI como número tipo 4188380701612.0
    - si es numérico -> str(int(...))
    - si es string -> trim
    """
    if v is None:
        return None
    if isinstance(v, (int,)):
        return str(v)
    if isinstance(v, (float,)):
        try:
            return str(int(v))
        except Exception:
            return _norm_str(v)
    s = str(v).strip()
    if s == "":
        return None
    # si viene "12345.0"
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


def _norm_lookup(s: str | None) -> str | None:
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    s = s.upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _cat_id_by_name(db: Session, table: str, name_col: str, value: str | None) -> int | None:
    """
    Lookup flexible (sin agregar columnas nuevas).
    Nota: comparamos UPPER(name_col) contra el valor normalizado (sin acentos).
    Si tus catálogos tienen 'nombre_normalizado', aquí puedes cambiar el WHERE.
    """
    v = _norm_lookup(value)
    if not v:
        return None

    row = db.execute(
        text(f"SELECT id FROM {table} WHERE UPPER({name_col}) = :v LIMIT 1"),
        {"v": v},
    ).scalar()
    return int(row) if row is not None else None


def _sexo_id(db: Session, value: str | None) -> int | None:
    s = _norm_lookup(value)
    if not s:
        return None
    if s in ("M", "MASCULINO", "HOMBRE", "1"):
        code = "M"
    elif s in ("F", "FEMENINO", "MUJER", "2"):
        code = "F"
    else:
        return None

    # intenta por codigo primero, luego por nombre
    sid = _cat_id_by_name(db, "cat_sexo", "codigo", code)
    if sid is None:
        sid = _cat_id_by_name(db, "cat_sexo", "nombre", code)
    return sid


def _validacion_id(db: Session, raw: str | None) -> int | None:
    """
    Si no mapea, devolvemos None -> core asigna INVALIDO automáticamente.
    """
    s = _norm_lookup(raw)
    if not s:
        return None
    # ejemplos típicos
    if s in ("VALIDO", "VÁLIDO"):
        vid = _cat_id_by_name(db, "cat_validacion", "codigo", "VALIDO")
        if vid is None:
            vid = _cat_id_by_name(db, "cat_validacion", "nombre", "VALIDO")
        return vid
    if s in ("INVALIDO", "INVÁLIDO"):
        vid = _cat_id_by_name(db, "cat_validacion", "codigo", "INVALIDO")
        if vid is None:
            vid = _cat_id_by_name(db, "cat_validacion", "nombre", "INVALIDO")
        return vid
    return None


def norm_header(s) -> str:
    """
    Normaliza headers para matching flexible:
    - trim
    - uppercase
    - quita acentos (AÑO->ANO, NIÑO->NINO)
    - deja A-Z0-9 y espacios
    - colapsa espacios
    """
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))  # quita acentos
    s = re.sub(r"[^A-Z0-9\s]", " ", s)  # quita símbolos, deja espacios
    s = re.sub(r"\s+", " ", s).strip()
    return s


def find_header_row(ws, max_scan_rows=40, max_scan_cols=80) -> int | None:
    """
    Busca la fila donde están los encabezados.
    Heurística: fila que contiene (normalizado) al menos estos “mínimos”:
      - CUI DEL NINO (o variantes)
      - NOMBRE DEL NINO (o variantes)
      - ANO / MES
    """
    triggers = [
        {"CUI DEL NINO", "CUI NINO", "CUI"},
        {"NOMBRE DEL NINO", "NOMBRE NINO"},
        {"ANO", "ANIO", "AÑO"},
        {"MES"},
    ]

    best_row = None
    best_score = -1

    for r in range(1, max_scan_rows + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_scan_cols + 1)]
        norm_vals = {norm_header(v) for v in row_vals if v not in (None, "")}

        score = 0
        for group in triggers:
            if any(g in norm_vals for g in group):
                score += 1

        if score > best_score:
            best_score = score
            best_row = r

        if score == len(triggers):
            return r

    if best_score >= 3:
        return best_row
    return None


def _read_sesan_xlsx_rows(file_bytes: bytes) -> list[dict]:
    """
    Lee el Excel SESAN en memoria.
    - Detecta la fila real del header automáticamente.
    - Normaliza headers (acentos, símbolos, dobles espacios).
    - No exige “25 columnas fijas”; solo exige mínimas.
    - Permite columnas extra (se guardan en raw_data).
    - Devuelve filas con keys canónicas para insertar staging.
    """
    bio = BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, data_only=True)

    ws = wb["SEVEROS"] if "SEVEROS" in wb.sheetnames else wb.active

    header_row = find_header_row(ws)
    if not header_row:
        raise HTTPException(
            status_code=422,
            detail="No se pudo detectar el encabezado del archivo SESAN (fila de títulos).",
        )

    max_cols = min(int(ws.max_column or 0) or 80, 120)
    raw_headers = [ws.cell(row=header_row, column=c).value for c in range(1, max_cols + 1)]
    norm_headers = [norm_header(h) for h in raw_headers]

    ALIASES = {
        "CIE 10": "CIE 10",
        "CIE-10": "CIE 10",
        "CIE_10": "CIE 10",
        "COMUNIDAD DE RESIDENCIA": "COMUNIDAD RESIDENCIA",
        "DIRECCION DE RESIDENCIA": "DIRECCION RESIDENCIA",
        "TELEFONOS DEL ENCARGADO": "TELEFONOS ENCARGADOS",
        "TELEFONO ENCARGADOS": "TELEFONOS ENCARGADOS",
        "TELEFONO DEL ENCARGADO": "TELEFONOS ENCARGADOS",
        "EDAD EN ANIOS": "EDAD EN ANOS",
        "EDAD EN AÑOS": "EDAD EN ANOS",
        "CUI DEL NIÑO": "CUI DEL NINO",
        "NOMBRE DEL NIÑO": "NOMBRE DEL NINO",
        "AÑO": "ANO",
        "ANIO": "ANO",
    }
    norm_headers = [ALIASES.get(h, h) for h in norm_headers]

    CANON = {
        "ANO": "ANO",
        "MES": "MES",
        "AREA DE SALUD": "AREA_DE_SALUD",
        "DISTRITO DE SALUD": "DISTRITO_DE_SALUD",
        "SERVICIO DE SALUD": "SERVICIO_DE_SALUD",
        "DEPARTAMENTO DE RESIDENCIA": "DEPTO_RESIDENCIA",
        "MUNICIPIO DE RESIDENCIA": "MUNI_RESIDENCIA",
        "COMUNIDAD RESIDENCIA": "COMUNIDAD_RESIDENCIA",
        "DIRECCION RESIDENCIA": "DIRECCION_RESIDENCIA",
        "CUI DEL NINO": "CUI_NINO",
        "SEXO": "SEXO",
        "EDAD EN ANOS": "EDAD_EN_ANOS",
        "NOMBRE DEL NINO": "NOMBRE_NINO",
        "FECHA NACIMIENTO": "FECHA_NACIMIENTO",
        "FECHA DEL PRIMER CONTACTO": "FECHA_PRIMER_CONTACTO",
        "FECHA DE REGISTRO": "FECHA_REGISTRO",
        "CIE 10": "CIE_10",
        "CIE_10": "CIE_10",
        "DIAGNOSTICO": "DIAGNOSTICO",
        "NOMBRE DE LA MADRE": "NOMBRE_MADRE",
        "CUI DE LA MADRE": "CUI_MADRE",
        "NOMBRE DEL PADRE": "NOMBRE_PADRE",
        "CUI DEL PADRE": "CUI_PADRE",
        "TELEFONOS ENCARGADOS": "TELEFONOS_ENCARGADOS",
        "VALIDACION": "VALIDACION",
    }

    present = {h for h in norm_headers if h}
    required_min = {"CUI DEL NINO", "NOMBRE DEL NINO", "ANO", "MES"}
    missing = sorted(required_min - present)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Estructura del archivo SESAN no coincide (faltan columnas críticas mínimas): {missing}",
        )

    col_to_key: dict[int, str] = {}
    for idx, h in enumerate(norm_headers, start=1):
        if h in CANON:
            col_to_key[idx] = CANON[h]

    rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        empty = 0
        row_canon: dict[str, object] = {}
        row_raw: dict[str, object] = {}

        for c in range(1, max_cols + 1):
            val = ws.cell(row=r, column=c).value
            hdr = norm_headers[c - 1] or f"COL_{c}"
            row_raw[hdr] = val

            if val is None or str(val).strip() == "":
                empty += 1

            key = col_to_key.get(c)
            if key:
                row_canon[key] = val

        if empty == max_cols:
            continue

        rows.append({"excel_row": r, "data": row_canon, "raw": row_raw})

    return rows


def _recalc_batch_counts(db: Session, batch_id: UUID):
    """
    Recalcula contadores y actualiza sesan_batch.
    También actualiza estado del batch:
      - si no hay pendientes -> FINALIZADO
      - si hay pendientes/errores/procesados -> EN_REVISION
      - si total=0 -> CARGADO
    """
    counts = db.execute(
        text("""
            SELECT
              COUNT(*) AS total,
              SUM(CASE WHEN estado = 'PENDIENTE' THEN 1 ELSE 0 END) AS pendientes,
              SUM(CASE WHEN estado = 'PROCESADO' THEN 1 ELSE 0 END) AS procesados,
              SUM(CASE WHEN estado = 'ERROR' THEN 1 ELSE 0 END) AS errores,
              SUM(CASE WHEN estado = 'IGNORADO' THEN 1 ELSE 0 END) AS ignorados
            FROM sesan_staging
            WHERE batch_id = :batch_id
        """),
        {"batch_id": str(batch_id)},
    ).mappings().one()

    total = int(counts["total"] or 0)
    pendientes = int(counts["pendientes"] or 0)
    procesados = int(counts["procesados"] or 0)
    errores = int(counts["errores"] or 0)
    ignorados = int(counts["ignorados"] or 0)

    if total <= 0:
        estado = "CARGADO"
    elif pendientes == 0:
        estado = "FINALIZADO"
    else:
        estado = "EN_REVISION"

    db.execute(
        text("""
            UPDATE sesan_batch
            SET
              total_registros = :total,
              total_pendientes = :pendientes,
              total_procesados = :procesados,
              total_error = :errores,
              total_ignorados = :ignorados,
              estado = :estado,
              updated_at = NOW()
            WHERE id = :batch_id
        """),
        {
            "batch_id": str(batch_id),
            "total": total,
            "pendientes": pendientes,
            "procesados": procesados,
            "errores": errores,
            "ignorados": ignorados,
            "estado": estado,
        },
    )


def _set_row_error(db: Session, row_id: UUID, code: str, msg: str):
    db.execute(
        text("""
            UPDATE sesan_staging
            SET
              estado = 'ERROR',
              error_code = :code,
              error_mensaje = :msg,
              intentos = COALESCE(intentos, 0) + 1,
              ultimo_intento_at = NOW(),
              updated_at = NOW()
            WHERE id = :id
        """),
        {"id": str(row_id), "code": code, "msg": msg},
    )


def _set_row_processed(db: Session, row_id: UUID, expediente_id: UUID):
    db.execute(
        text("""
            UPDATE sesan_staging
            SET
              estado = 'PROCESADO',
              expediente_id = :expediente_id,
              error_code = NULL,
              error_mensaje = NULL,
              intentos = COALESCE(intentos, 0) + 1,
              ultimo_intento_at = NOW(),
              updated_at = NOW()
            WHERE id = :id
        """),
        {"id": str(row_id), "expediente_id": str(expediente_id)},
    )


def _is_dup_cui_in_year(db: Session, cui_nino: str, anio_carga: int, current_row_id: UUID) -> bool:
    """
    Regla: No se puede repetir CUI por año de carga (anio_carga viene del batch).
    Validación contra staging ya PROCESADO en cualquier batch del mismo anio_carga.
    """
    exists = db.execute(
        text("""
            SELECT 1
            FROM sesan_staging s
            JOIN sesan_batch b ON b.id = s.batch_id
            WHERE b.anio_carga = :anio
              AND s.estado = 'PROCESADO'
              AND s.cui_nino = :cui
              AND s.id <> :row_id
            LIMIT 1
        """),
        {"anio": anio_carga, "cui": cui_nino, "row_id": str(current_row_id)},
    ).scalar()
    return bool(exists)


def _is_dup_cui_in_expedientes(db: Session, cui_nino: str, anio_carga: int) -> bool:
    """
    Refuerzo de regla: valida también contra expedientes ya creados (info_general).
    Usa info_general.cui_del_nino y info_general.anio (string) como referencia.
    """
    exists = db.execute(
        text("""
            SELECT 1
            FROM info_general ig
            WHERE ig.cui_del_nino = :cui
              AND ig.anio = :anio
            LIMIT 1
        """),
        {"cui": cui_nino, "anio": str(anio_carga)},
    ).scalar()
    return bool(exists)


def _build_expediente_payload_from_row(db: Session, row: dict, anio_carga: int, mes_carga: int | None):
    """
    Normaliza y mapea sesan_staging -> ExpedienteCreate (con InfoGeneralIn)
    """
    cui_nino = _to_cui(row.get("cui_nino"))
    nombre_nino = _norm_str(row.get("nombre_nino"))

    # Territorio (residencia)
    depto_res_id = _cat_id_by_name(db, "cat_departamento", "nombre", row.get("departamento_residencia"))
    muni_res_id = _cat_id_by_name(db, "cat_municipio", "nombre", row.get("municipio_residencia"))

    # Salud
    area_id = _cat_id_by_name(db, "cat_area_salud", "nombre", row.get("area_salud"))
    distrito_id = _cat_id_by_name(db, "cat_distrito_salud", "nombre", row.get("distrito_salud"))
    servicio_id = _cat_id_by_name(db, "cat_servicio_salud", "nombre", row.get("servicio_salud"))

    sexo_id = _sexo_id(db, row.get("sexo"))
    validacion_id = _validacion_id(db, row.get("validacion_raw"))  # si None -> core pone INVALIDO

    # Usamos anio_carga como “anio” de MIS (regla de negocio)
    ig_anio = str(anio_carga)
    ig_mes = str(_to_int(row.get("mes")) or mes_carga or "") or None

    ig = InfoGeneralIn(
        anio=ig_anio,
        mes=ig_mes,

        area_salud_id=area_id,
        distrito_salud_id=distrito_id,
        servicio_salud_id=servicio_id,

        departamento_residencia_id=depto_res_id,
        municipio_residencia_id=muni_res_id,
        comunidad_residencia=_norm_str(row.get("comunidad_residencia")),
        direccion_residencia=_norm_str(row.get("direccion_residencia")),

        cui_del_nino=cui_nino,
        sexo_id=sexo_id,
        edad_en_anios=_norm_str(row.get("edad_en_anios")),
        nombre_del_nino=nombre_nino,

        fecha_nacimiento=row.get("fecha_nacimiento"),
        fecha_del_primer_contacto=row.get("fecha_primer_contacto"),
        fecha_de_registro=row.get("fecha_registro"),

        cie_10=_norm_str(row.get("cie_10")),
        diagnostico=_norm_str(row.get("diagnostico")),

        nombre_de_la_madre=_norm_str(row.get("nombre_madre")),
        cui_de_la_madre=_to_cui(row.get("cui_madre")),
        nombre_del_padre=_norm_str(row.get("nombre_padre")),
        cui_del_padre=_to_cui(row.get("cui_padre")),

        telefonos_encargados=_norm_str(row.get("telefonos_encargados")),

        validacion_id=validacion_id,
    )

    # Expediente “principal”
    payload = ExpedienteCreate(
        nombre_beneficiario=nombre_nino,
        cui_beneficiario=cui_nino,
        departamento_id=depto_res_id,
        municipio_id=muni_res_id,
        info_general=ig,
    )

    return payload


def _procesar_row_creando_expediente(db: Session, row_id: UUID):
    """
    Procesa 1 fila:
    - valida mínimos
    - regla no duplicar
    - normaliza + crea expediente
    - marca PROCESADO con expediente_id
    """
    row = db.execute(
        text("""
            SELECT
              s.*,
              b.anio_carga,
              b.mes_carga
            FROM sesan_staging s
            JOIN sesan_batch b ON b.id = s.batch_id
            WHERE s.id = :id
            FOR UPDATE
        """),
        {"id": str(row_id)},
    ).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="Fila staging no encontrada.")

    if row["estado"] == "IGNORADO":
        raise HTTPException(status_code=409, detail="La fila está IGNORADA.")

    if row["estado"] == "PROCESADO" and row.get("expediente_id"):
        return {
            "row_id": str(row_id),
            "estado": "PROCESADO",
            "expediente_id": str(row["expediente_id"]),
        }

    anio_carga = int(row["anio_carga"])
    mes_carga = int(row["mes_carga"]) if row.get("mes_carga") is not None else None

    cui = _to_cui(row.get("cui_nino"))
    nombre = _norm_str(row.get("nombre_nino"))

    if not cui:
        raise ValueError("MISSING_CUI|CUI del niño vacío.")
    if not nombre:
        raise ValueError("MISSING_NAME|Nombre del niño vacío.")

    if _is_dup_cui_in_year(db, cui, anio_carga, row_id):
        raise ValueError(f"DUP_CUI_YEAR|CUI duplicado en el año de carga {anio_carga} (staging).")

    if _is_dup_cui_in_expedientes(db, cui, anio_carga):
        raise ValueError(f"DUP_CUI_YEAR|CUI duplicado en el año de carga {anio_carga} (expedientes).")

    # construir payload ya normalizado
    payload = _build_expediente_payload_from_row(db, row, anio_carga, mes_carga)

    # ✅ crear expediente usando el CORE oficial
    exp = crear_expediente_core(payload, db)

    # marcar fila procesada con expediente_id
    _set_row_processed(db, row_id, exp.id)

    # recalcular batch
    _recalc_batch_counts(db, UUID(str(row["batch_id"])))

    return {
        "row_id": str(row_id),
        "estado": "PROCESADO",
        "expediente_id": str(exp.id),
    }


# =====================================================
# 1) Crear batch + staging (SUBIDA)
# =====================================================
@router.post("/batch", status_code=201)
def crear_batch_sesan(
    db: Session = Depends(get_db),

    # Metadata
    nombre_lote: str = Form(...),
    anio_carga: int = Form(...),
    mes_carga: int | None = Form(None),
    descripcion: str | None = Form(None),
    origen: str = Form("SESAN"),
    usuario_carga: str | None = Form(None),

    # Archivo
    file: UploadFile = File(...),
):
    """
    - Lee el archivo en memoria
    - Guarda referencia simulada (FTP)
    - Crea sesan_batch
    - Inserta sesan_staging (PENDIENTE)
    """
    try:
        file_bytes = file.file.read()
        if not file_bytes:
            raise HTTPException(status_code=400, detail="Archivo vacío.")

        size_bytes = len(file_bytes)
        checksum = _sha256_bytes(file_bytes)

        ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        safe_name = (file.filename or "sesan.xlsx").replace("\\", "_").replace("/", "_")
        storage_provider = "ftp"
        storage_key = f"ftp://PENDIENTE/sesan/{ts}_{safe_name}"

        batch_id = db.execute(
            text("""
                INSERT INTO sesan_batch (
                  nombre_lote, descripcion, origen,
                  anio_carga, mes_carga, usuario_carga,
                  archivo_nombre_original, archivo_mime_type, archivo_size_bytes,
                  storage_provider, storage_key, checksum_sha256,
                  estado,
                  total_registros, total_pendientes, total_procesados, total_error, total_ignorados,
                  created_at, updated_at
                )
                VALUES (
                  :nombre_lote, :descripcion, :origen,
                  :anio_carga, :mes_carga, :usuario_carga,
                  :archivo_nombre_original, :archivo_mime_type, :archivo_size_bytes,
                  :storage_provider, :storage_key, :checksum_sha256,
                  'CARGADO',
                  0, 0, 0, 0, 0,
                  NOW(), NOW()
                )
                RETURNING id
            """),
            {
                "nombre_lote": nombre_lote,
                "descripcion": descripcion,
                "origen": origen,
                "anio_carga": anio_carga,
                "mes_carga": mes_carga,
                "usuario_carga": usuario_carga,
                "archivo_nombre_original": file.filename or "sesan.xlsx",
                "archivo_mime_type": file.content_type,
                "archivo_size_bytes": size_bytes,
                "storage_provider": storage_provider,
                "storage_key": storage_key,
                "checksum_sha256": checksum,
            }
        ).scalar_one()

        rows = _read_sesan_xlsx_rows(file_bytes)
        if not rows:
            raise HTTPException(status_code=422, detail="No se encontraron filas válidas.")

        insert_staging = text("""
            INSERT INTO sesan_staging (
              batch_id, row_num,
              anio, mes, area_salud, distrito_salud, servicio_salud,
              departamento_residencia, municipio_residencia, comunidad_residencia, direccion_residencia,
              cui_nino, sexo, edad_en_anios, nombre_nino,
              fecha_nacimiento, fecha_primer_contacto, fecha_registro,
              cie_10, diagnostico,
              nombre_madre, cui_madre, nombre_padre, cui_padre, telefonos_encargados,
              validacion_raw,
              raw_data,
              estado,
              created_at, updated_at
            )
            VALUES (
              :batch_id, :row_num,
              :anio, :mes, :area_salud, :distrito_salud, :servicio_salud,
              :departamento_residencia, :municipio_residencia, :comunidad_residencia, :direccion_residencia,
              :cui_nino, :sexo, :edad_en_anios, :nombre_nino,
              :fecha_nacimiento, :fecha_primer_contacto, :fecha_registro,
              :cie_10, :diagnostico,
              :nombre_madre, :cui_madre, :nombre_padre, :cui_padre, :telefonos_encargados,
              :validacion_raw,
              CAST(:raw_data AS jsonb),
              'PENDIENTE',
              NOW(), NOW()
            )
        """)

        total = 0
        for item in rows:
            r = item["data"]
            raw_for_audit = item.get("raw") or {}

            db.execute(
                insert_staging,
                {
                    "batch_id": str(batch_id),
                    "row_num": item["excel_row"],

                    "anio": _to_int(r.get("ANO")),
                    "mes": _to_int(r.get("MES")),
                    "area_salud": _norm_str(r.get("AREA_DE_SALUD")),
                    "distrito_salud": _norm_str(r.get("DISTRITO_DE_SALUD")),
                    "servicio_salud": _norm_str(r.get("SERVICIO_DE_SALUD")),

                    "departamento_residencia": _norm_str(r.get("DEPTO_RESIDENCIA")),
                    "municipio_residencia": _norm_str(r.get("MUNI_RESIDENCIA")),
                    "comunidad_residencia": _norm_str(r.get("COMUNIDAD_RESIDENCIA")),
                    "direccion_residencia": _norm_str(r.get("DIRECCION_RESIDENCIA")),

                    "cui_nino": _to_cui(r.get("CUI_NINO")),
                    "sexo": _norm_str(r.get("SEXO")),
                    "edad_en_anios": _norm_str(r.get("EDAD_EN_ANOS")),
                    "nombre_nino": _norm_str(r.get("NOMBRE_NINO")),

                    "fecha_nacimiento": _to_date(r.get("FECHA_NACIMIENTO")),
                    "fecha_primer_contacto": _to_date(r.get("FECHA_PRIMER_CONTACTO")),
                    "fecha_registro": _to_date(r.get("FECHA_REGISTRO")),

                    "cie_10": _norm_str(r.get("CIE_10")),
                    "diagnostico": _norm_str(r.get("DIAGNOSTICO")),

                    "nombre_madre": _norm_str(r.get("NOMBRE_MADRE")),
                    "cui_madre": _to_cui(r.get("CUI_MADRE")),
                    "nombre_padre": _norm_str(r.get("NOMBRE_PADRE")),
                    "cui_padre": _to_cui(r.get("CUI_PADRE")),
                    "telefonos_encargados": _norm_str(r.get("TELEFONOS_ENCARGADOS")),

                    "validacion_raw": _norm_str(r.get("VALIDACION")),

                    "raw_data": json.dumps(raw_for_audit, default=str),
                }
            )
            total += 1

        _recalc_batch_counts(db, UUID(str(batch_id)))
        db.commit()

        return {
            "batch_id": str(batch_id),
            "total_registros": total,
            "storage_key": storage_key,
            "checksum_sha256": checksum,
        }

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creando batch SESAN: {str(e)}")


# =====================================================
# 2) Listar batches por año
# =====================================================
@router.get("/batches")
def listar_batches_por_anio(
    anio: int = Query(...),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    offset = (page - 1) * limit

    total = db.execute(
        text("SELECT COUNT(*) FROM sesan_batch WHERE anio_carga = :anio"),
        {"anio": anio},
    ).scalar() or 0

    rows = db.execute(
        text("""
            SELECT *
            FROM sesan_batch
            WHERE anio_carga = :anio
            ORDER BY created_at DESC
            OFFSET :offset
            LIMIT :limit
        """),
        {"anio": anio, "offset": offset, "limit": limit},
    ).mappings().all()

    return {
        "page": page,
        "limit": limit,
        "total": total,
        "data": [dict(r) for r in rows],
    }


# =====================================================
# 3) Listar filas por batch
# =====================================================
@router.get("/batch/{batch_id}/rows")
def listar_filas_batch(
    batch_id: UUID,
    estado: str | None = Query(None, description="PENDIENTE | ERROR | PROCESADO | IGNORADO"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    offset = (page - 1) * limit

    base = "FROM sesan_staging WHERE batch_id = :batch_id"
    params = {"batch_id": str(batch_id)}

    if estado:
        base += " AND estado = :estado"
        params["estado"] = estado

    total = db.execute(
        text(f"SELECT COUNT(*) {base}"),
        params,
    ).scalar() or 0

    rows = db.execute(
        text(f"""
            SELECT
              id, row_num, estado, error_code, error_mensaje,
              cui_nino, nombre_nino,
              departamento_residencia, municipio_residencia,
              cie_10, diagnostico,
              expediente_id,
              intentos, ultimo_intento_at,
              corregido_por, corregido_at,
              ignorado_por, ignorado_at, motivo_ignorado
            {base}
            ORDER BY row_num ASC
            OFFSET :offset
            LIMIT :limit
        """),
        {**params, "offset": offset, "limit": limit},
    ).mappings().all()

    return {
        "page": page,
        "limit": limit,
        "total": total,
        "data": [dict(r) for r in rows],
    }


@router.get("/anios")
def listar_anios_sesan(db: Session = Depends(get_db)):
    rows = db.execute(
        text("""
            SELECT
              anio_carga,
              COUNT(*) AS total_batches
            FROM sesan_batch
            GROUP BY anio_carga
            ORDER BY anio_carga DESC
        """)
    ).mappings().all()

    return {
        "data": [
            {"anio_carga": r["anio_carga"], "total_batches": r["total_batches"]}
            for r in rows
        ]
    }


# =====================================================
# 4) Procesamiento / Reintentos / Ignorar
# =====================================================

@router.post("/batch/{batch_id}/procesar-pendientes")
def procesar_pendientes_batch(
    batch_id: UUID,
    limit: int = Query(200, ge=1, le=2000),
    db: Session = Depends(get_db),
):
    """
    Procesa filas PENDIENTE del batch (hasta limit) creando expediente electrónico.
    """
    try:
        batch = db.execute(
            text("SELECT id, anio_carga FROM sesan_batch WHERE id = :id"),
            {"id": str(batch_id)},
        ).mappings().first()

        if not batch:
            raise HTTPException(status_code=404, detail="Batch no encontrado.")

        rows = db.execute(
            text("""
                SELECT id
                FROM sesan_staging
                WHERE batch_id = :batch_id
                  AND estado = 'PENDIENTE'
                ORDER BY row_num ASC
                LIMIT :limit
            """),
            {"batch_id": str(batch_id), "limit": limit},
        ).mappings().all()

        procesados = 0
        errores = 0

        for r in rows:
            rid = UUID(str(r["id"]))
            try:
                _procesar_row_creando_expediente(db, rid)
                procesados += 1
            except ValueError as ve:
                # formato: CODE|Mensaje
                raw = str(ve)
                if "|" in raw:
                    code, msg = raw.split("|", 1)
                else:
                    code, msg = "VALIDATION_ERROR", raw
                _set_row_error(db, rid, code.strip(), msg.strip())
                errores += 1
            except HTTPException as he:
                _set_row_error(db, rid, "HTTP_ERROR", str(he.detail))
                errores += 1
            except Exception as e:
                _set_row_error(db, rid, "UNEXPECTED_ERROR", str(e))
                errores += 1

        _recalc_batch_counts(db, batch_id)
        db.commit()

        return {
            "batch_id": str(batch_id),
            "procesados": procesados,
            "errores": errores,
            "total_intentados": len(rows),
        }

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error procesando pendientes: {str(e)}")


@router.post("/row/{row_id}/procesar")
def procesar_row(
    row_id: UUID,
    db: Session = Depends(get_db),
):
    """
    Procesa 1 fila creando expediente.
    Ideal para el botón individual por registro.
    """
    try:
        result = _procesar_row_creando_expediente(db, row_id)
        db.commit()
        return result
    except HTTPException:
        db.rollback()
        raise
    except ValueError as ve:
        db.rollback()
        raw = str(ve)
        if "|" in raw:
            code, msg = raw.split("|", 1)
        else:
            code, msg = "VALIDATION_ERROR", raw
        _set_row_error(db, row_id, code.strip(), msg.strip())
        # recalcular batch (si existe)
        try:
            b = db.execute(text("SELECT batch_id FROM sesan_staging WHERE id=:id"), {"id": str(row_id)}).scalar()
            if b:
                _recalc_batch_counts(db, UUID(str(b)))
            db.commit()
        except Exception:
            db.rollback()
        raise HTTPException(status_code=422, detail=msg.strip())
    except Exception as e:
        db.rollback()
        _set_row_error(db, row_id, "UNEXPECTED_ERROR", str(e))
        try:
            b = db.execute(text("SELECT batch_id FROM sesan_staging WHERE id=:id"), {"id": str(row_id)}).scalar()
            if b:
                _recalc_batch_counts(db, UUID(str(b)))
            db.commit()
        except Exception:
            db.rollback()
        raise HTTPException(status_code=500, detail=f"Error procesando fila: {str(e)}")


@router.post("/batch/{batch_id}/reintentar-errores")
def reintentar_errores_batch(
    batch_id: UUID,
    limit: int = Query(2000, ge=1, le=50000),
    db: Session = Depends(get_db),
):
    """
    Pasa filas ERROR -> PENDIENTE (para re-procesar).
    ✅ Corregido: Postgres no permite LIMIT directo en UPDATE; usamos CTE.
    """
    try:
        updated = db.execute(
            text("""
                WITH to_update AS (
                  SELECT id
                  FROM sesan_staging
                  WHERE batch_id = :batch_id
                    AND estado = 'ERROR'
                  ORDER BY row_num ASC
                  LIMIT :limit
                )
                UPDATE sesan_staging s
                SET
                  estado = 'PENDIENTE',
                  error_code = NULL,
                  error_mensaje = NULL,
                  updated_at = NOW()
                FROM to_update u
                WHERE s.id = u.id
                RETURNING s.id
            """),
            {"batch_id": str(batch_id), "limit": limit},
        ).fetchall()

        _recalc_batch_counts(db, batch_id)
        db.commit()

        return {
            "batch_id": str(batch_id),
            "rows_reintentadas": len(updated),
        }

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error reintentando errores: {str(e)}")


@router.post("/row/{row_id}/reintentar")
def reintentar_row(
    row_id: UUID,
    db: Session = Depends(get_db),
):
    """
    Pasa una fila a PENDIENTE (limpia error) para reintentar.
    """
    try:
        row = db.execute(
            text("SELECT id, batch_id FROM sesan_staging WHERE id = :id"),
            {"id": str(row_id)},
        ).mappings().first()

        if not row:
            raise HTTPException(status_code=404, detail="Fila staging no encontrada.")

        db.execute(
            text("""
                UPDATE sesan_staging
                SET
                  estado = 'PENDIENTE',
                  error_code = NULL,
                  error_mensaje = NULL,
                  updated_at = NOW()
                WHERE id = :id
            """),
            {"id": str(row_id)},
        )

        _recalc_batch_counts(db, UUID(str(row["batch_id"])))
        db.commit()

        return {"row_id": str(row_id), "estado": "PENDIENTE"}

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error reintentando fila: {str(e)}")


@router.post("/row/{row_id}/ignorar")
def ignorar_row(
    row_id: UUID,
    motivo: str = Form(...),
    usuario: str | None = Form(None),
    db: Session = Depends(get_db),
):
    """
    Marca una fila como IGNORADO (con motivo).
    """
    try:
        row = db.execute(
            text("SELECT id, batch_id FROM sesan_staging WHERE id = :id"),
            {"id": str(row_id)},
        ).mappings().first()

        if not row:
            raise HTTPException(status_code=404, detail="Fila staging no encontrada.")

        db.execute(
            text("""
                UPDATE sesan_staging
                SET
                  estado = 'IGNORADO',
                  motivo_ignorado = :motivo,
                  ignorado_por = :usuario,
                  ignorado_at = NOW(),
                  updated_at = NOW()
                WHERE id = :id
            """),
            {"id": str(row_id), "motivo": motivo, "usuario": usuario},
        )

        _recalc_batch_counts(db, UUID(str(row["batch_id"])))
        db.commit()

        return {"row_id": str(row_id), "estado": "IGNORADO"}

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error ignorando fila: {str(e)}")
