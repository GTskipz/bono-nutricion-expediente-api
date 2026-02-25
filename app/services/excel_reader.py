from __future__ import annotations

from fastapi import HTTPException
import openpyxl
import re
import unicodedata
from typing import Iterator, Optional


def norm_header(s) -> str:
    """
    Normaliza headers para matching flexible:
    - trim
    - uppercase
    - quita acentos (AÑO->ANO, NIÑO->NINO)
    - deja A-Z0-9 y espacios
    - colapsa espacios

    ✅ Caso especial: si el header literal es "#", significa RUB.
    """
    if s is None:
        return ""
    raw = str(s).strip()

    if raw == "#":
        return "RUB"

    s = raw.upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# =====================================================
# ✅ NUEVO: find_header_row compatible con read_only (iter_rows)
# =====================================================
def find_header_row_iter(ws, max_scan_rows: int = 40, max_scan_cols: int = 80) -> int | None:
    """
    Busca la fila donde están los encabezados.
    Heurística: fila que contiene (normalizado) al menos estos “mínimos”:
      - CUI DEL NINO (o variantes)
      - NOMBRE DEL NINO (o variantes)
      - ANO / MES

    ✅ Compatible con openpyxl read_only=True usando iter_rows(values_only=True)
    """
    triggers = [
        {"CUI DEL NINO", "CUI NINO", "CUI"},
        {"NOMBRE DEL NINO", "NOMBRE NINO"},
        {"ANO", "ANIO", "AÑO"},
        {"MES"},
    ]

    best_row = None
    best_score = -1

    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, max_col=max_scan_cols, values_only=True),
        start=1
    ):
        norm_vals = {norm_header(v) for v in row if v not in (None, "")}

        score = 0
        for group in triggers:
            if any(g in norm_vals for g in group):
                score += 1

        if score > best_score:
            best_score = score
            best_row = idx

        if score == len(triggers):
            return idx

    if best_score >= 3:
        return best_row
    return None


# =====================================================
# ✅ NUEVO: iterator streaming (NO lista gigante)
# =====================================================
def iter_sesan_xlsx_rows(file_path: str) -> Iterator[dict]:
    """
    Lee el Excel SESAN en streaming desde disco.
    - Detecta fila de header automáticamente (read_only)
    - Normaliza headers
    - Permite columnas extra (raw)
    - Devuelve filas con keys canónicas para staging
    - NO construye lista: yield fila por fila ✅ (robusto 60k+)

    Uso:
        for item in iter_sesan_xlsx_rows(tmp_path):
            ...
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb["SEVEROS"] if "SEVEROS" in wb.sheetnames else wb.active

    header_row = find_header_row_iter(ws)
    if not header_row:
        raise HTTPException(
            status_code=422,
            detail="No se pudo detectar el encabezado del archivo SESAN (fila de títulos).",
        )

    # Leer headers (fila header_row)
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), None)
    if not header_values:
        raise HTTPException(status_code=422, detail="No se pudo leer el encabezado del archivo SESAN.")

    raw_headers = list(header_values)
    # Limitar columnas (en read_only max_column no es confiable)
    max_cols = min(len(raw_headers) if len(raw_headers) > 0 else 80, 120)
    raw_headers = raw_headers[:max_cols]

    norm_headers = [norm_header(h) for h in raw_headers]

    ALIASES = {
        "CIE 10": "CIE 10",
        "CIE-10": "CIE 10",
        "CIE_10": "CIE 10",
        "COMUNIDAD DE RESIDENCIA": "COMUNIDAD RESIDENCIA",
        "DIRECCION DE RESIDENCIA": "DIRECCION RESIDENCIA",
        "TELEFONOS DEL ENCARGADO": "TELEFONOS ENCARGADOS",
        "TELEFONO ENCARGADOS": "TELEFONOS ENCARGADOS",
        "TELEFONO DEL ENCARGADO": "TELEFONOS ENCARGADOS",
        "EDAD EN ANIOS": "EDAD EN ANOS",
        "EDAD EN AÑOS": "EDAD EN ANOS",
        "CUI DEL NIÑO": "CUI DEL NINO",
        "NOMBRE DEL NIÑO": "NOMBRE DEL NINO",
        "AÑO": "ANO",
        "ANIO": "ANO",
        "REGISTRO UNICO DE BENEFICIARIO": "RUB",
        "REGISTRO ÚNICO DE BENEFICIARIO": "RUB",
        "REGISTRO UNICO BENEFICIARIO": "RUB",
        "REGISTRO ÚNICO BENEFICIARIO": "RUB",
    }
    norm_headers = [ALIASES.get(h, h) for h in norm_headers]

    CANON = {
        "RUB": "RUB",
        "ANO": "ANO",
        "MES": "MES",
        "AREA DE SALUD": "AREA_DE_SALUD",
        "DISTRITO DE SALUD": "DISTRITO_DE_SALUD",
        "SERVICIO DE SALUD": "SERVICIO_DE_SALUD",
        "DEPARTAMENTO DE RESIDENCIA": "DEPTO_RESIDENCIA",
        "MUNICIPIO DE RESIDENCIA": "MUNI_RESIDENCIA",
        "COMUNIDAD RESIDENCIA": "COMUNIDAD_RESIDENCIA",
        "DIRECCION RESIDENCIA": "DIRECCION_RESIDENCIA",
        "CUI DEL NINO": "CUI_NINO",
        "SEXO": "SEXO",
        "EDAD EN ANOS": "EDAD_EN_ANOS",
        "NOMBRE DEL NINO": "NOMBRE_NINO",
        "FECHA NACIMIENTO": "FECHA_NACIMIENTO",
        "FECHA DEL PRIMER CONTACTO": "FECHA_PRIMER_CONTACTO",
        "FECHA DE REGISTRO": "FECHA_REGISTRO",
        "CIE 10": "CIE_10",
        "CIE_10": "CIE_10",
        "DIAGNOSTICO": "DIAGNOSTICO",
        "NOMBRE DE LA MADRE": "NOMBRE_MADRE",
        "CUI DE LA MADRE": "CUI_MADRE",
        "NOMBRE DEL PADRE": "NOMBRE_PADRE",
        "CUI DEL PADRE": "CUI_PADRE",
        "TELEFONOS ENCARGADOS": "TELEFONOS_ENCARGADOS",
        "VALIDACION": "VALIDACION",
    }

    present = {h for h in norm_headers if h}
    required_min = {"CUI DEL NINO", "NOMBRE DEL NINO", "ANO", "MES"}  # RUB opcional
    missing = sorted(required_min - present)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Estructura del archivo SESAN no coincide (faltan columnas críticas mínimas): {missing}",
        )

    # mapa col (1-based) -> key canon
    col_to_key: dict[int, str] = {}
    for idx, h in enumerate(norm_headers, start=1):
        if h in CANON:
            col_to_key[idx] = CANON[h]

    # Iterar filas de datos
    excel_row = header_row
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        excel_row += 1
        row = tuple(row[:max_cols])  # limitar

        if all(v is None or str(v).strip() == "" for v in row):
            continue

        row_canon: dict[str, object] = {}
        row_raw: dict[str, object] = {}

        for c in range(1, max_cols + 1):
            val = row[c - 1] if (c - 1) < len(row) else None
            hdr = norm_headers[c - 1] or f"COL_{c}"
            row_raw[hdr] = val

            key = col_to_key.get(c)
            if key:
                row_canon[key] = val

        yield {"excel_row": excel_row, "data": row_canon, "raw": row_raw}


# =====================================================
# ✅ Mantener API vieja (compatibilidad): retorna lista
#    pero ya no lee bytes: recibe file_path y convierte el iterador en lista
# =====================================================
def read_sesan_xlsx_rows(file_path: str) -> list[dict]:
    """
    Compatibilidad: ahora lee desde file_path y arma lista.
    ⚠️ Para archivos grandes, usa iter_sesan_xlsx_rows() directamente.
    """
    return list(iter_sesan_xlsx_rows(file_path))
